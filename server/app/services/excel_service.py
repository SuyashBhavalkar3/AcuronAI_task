import io
import logging
from typing import List

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.schemas.invoice import InvoiceProcessingResult

logger = logging.getLogger(__name__)

COLUMN_HEADERS = [
    "Acc Period", "Trans Date", "Account Code", "Curr Code", "Trans Amount",
    "Dr_Cr", "Jrnal Type", "Jrnal Source", "Reference", "Description",
    "Asset Code", "Asset Indicator", "Asset / Item Qty", "Due Date",
    "Branch Analysis Code", "Product Analysis Code", "ChannelAnalysisCode",
    "Sub-Channel Analysis Code", "Underwriting Year Analysis Code",
    "Employee Code Analysis Code", "TDS Applicability Analysis Code",
    "Department Analysis Code", "Sequence Code Analysis Code",
    "Vendor Code Analysis Code", "Invoice Date", "From Date", "To Date",
    "Addl Date 4", "Addl Date 5", "Cheque & NEFT Number", "Invoice Number",
    "Additional Remarks", "Additional Remarks 2", "CREDENCE DESCRIPTION",
    "HSN/SAC NO", "Taxable on Amount", "Reverse Charge (Y/N)",
    "Reverse charge %", "Item Details (Sr.No)", "Goods/Service",
    "GST Tax Rate", "Orginal Invoice no for Dr/Cr Notes", "Advance Challan No",
]


def _row_to_dict(result: InvoiceProcessingResult) -> dict:
    """Convert an InvoiceProcessingResult accounting_row to a flat dict."""
    r = result.accounting_row
    if r is None:
        return {col: None for col in COLUMN_HEADERS}
    return {
        "Acc Period": r.acc_period,
        "Trans Date": r.trans_date,
        "Account Code": r.account_code,
        "Curr Code": r.curr_code,
        "Trans Amount": r.trans_amount,
        "Dr_Cr": r.dr_cr,
        "Jrnal Type": r.jrnal_type,
        "Jrnal Source": r.jrnal_source,
        "Reference": r.reference,
        "Description": r.description,
        "Asset Code": r.asset_code,
        "Asset Indicator": r.asset_indicator,
        "Asset / Item Qty": r.asset_item_qty,
        "Due Date": r.due_date,
        "Branch Analysis Code": r.branch_analysis_code,
        "Product Analysis Code": r.product_analysis_code,
        "ChannelAnalysisCode": r.channel_analysis_code,
        "Sub-Channel Analysis Code": r.sub_channel_analysis_code,
        "Underwriting Year Analysis Code": r.underwriting_year_analysis_code,
        "Employee Code Analysis Code": r.employee_code_analysis_code,
        "TDS Applicability Analysis Code": r.tds_applicability_analysis_code,
        "Department Analysis Code": r.department_analysis_code,
        "Sequence Code Analysis Code": r.sequence_code_analysis_code,
        "Vendor Code Analysis Code": r.vendor_code_analysis_code,
        "Invoice Date": r.invoice_date,
        "From Date": r.from_date,
        "To Date": r.to_date,
        "Addl Date 4": r.addl_date_4,
        "Addl Date 5": r.addl_date_5,
        "Cheque & NEFT Number": r.cheque_neft_number,
        "Invoice Number": r.invoice_number,
        "Additional Remarks": r.additional_remarks,
        "Additional Remarks 2": r.additional_remarks_2,
        "CREDENCE DESCRIPTION": r.credence_description,
        "HSN/SAC NO": r.hsn_sac_no,
        "Taxable on Amount": r.taxable_on_amount,
        "Reverse Charge (Y/N)": r.reverse_charge,
        "Reverse charge %": r.reverse_charge_pct,
        "Item Details (Sr.No)": r.item_details_sr_no,
        "Goods/Service": r.goods_service,
        "GST Tax Rate": r.gst_tax_rate,
        "Orginal Invoice no for Dr/Cr Notes": r.original_invoice_no_dr_cr,
        "Advance Challan No": r.advance_challan_no,
    }


def generate_excel(results: List[InvoiceProcessingResult]) -> bytes:
    """
    Generate a styled Excel workbook from processing results and return bytes.
    """
    wb = Workbook()

    # ── Sheet 1: Accounting Data ──────────────────────────────────────────────
    ws = wb.active
    ws.title = "Invoice Accounting"

    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    success_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    warning_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    error_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")

    # Add filename column first
    all_headers = ["File Name", "Status"] + COLUMN_HEADERS
    for col_idx, header in enumerate(all_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "C2"

    row_idx = 2
    for result in results:
        e = result.extracted
        row_fill = success_fill
        if result.status == "error":
            row_fill = error_fill
        elif result.status == "warning":
            row_fill = warning_fill

        # Get tax breakdown or create a dummy one if none exists
        breakdown = []
        if e and e.tax_breakdown:
            breakdown = e.tax_breakdown
        else:
            # Fallback: create a single entry from the top-level data
            from app.schemas.invoice import TaxRateBreakdown
            breakdown = [TaxRateBreakdown(
                rate=float(str(e.gst_rate).replace("%", "").split(",")[0]) if e and e.gst_rate else 0,
                taxable_amount=e.taxable_amount if e else 0,
                gst_amount=e.gst_amount if e else 0
            )]

        for entry in breakdown:
            row_data = _row_to_dict(result)
            # Update the split-specific fields
            row_data["Taxable on Amount"] = entry.taxable_amount
            row_data["GST Tax Rate"] = f"{entry.rate}%"
            # If trans_amount should be the line total (taxable + gst for that rate)
            row_data["Trans Amount"] = round(entry.taxable_amount + entry.gst_amount, 2)

            # File name and status
            ws.cell(row=row_idx, column=1, value=result.filename).fill = row_fill
            ws.cell(row=row_idx, column=2, value=result.status.upper()).fill = row_fill

            for col_idx, header in enumerate(COLUMN_HEADERS, start=3):
                cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(header))
                cell.fill = row_fill
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")
            
            row_idx += 1

    # Auto-size columns (cap at 40)
    for col_idx in range(1, len(all_headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            (len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, ws.max_row + 1)),
            default=10,
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    # ── Sheet 2: Full Extracted Data ──────────────────────────────────────────
    ws2 = wb.create_sheet("Full Extracted Data")
    ext_headers = [
        "File Name", "Vendor Name", "Vendor GSTIN", "Invoice Number", 
        "Invoice Date", "Taxable Amount", "GST Amount", "GST Rate", 
        "HSN/SAC", "Total Amount"
    ]
    for col_idx, h in enumerate(ext_headers, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for row_idx, result in enumerate(results, start=2):
        e = result.extracted
        if not e:
            continue
        vals = [
            result.filename, e.vendor_name, e.vendor_gstin, e.invoice_number,
            e.invoice_date, e.taxable_amount, e.gst_amount, e.gst_rate,
            e.hsn_sac, e.total_amount
        ]
        for col_idx, val in enumerate(vals, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            
    for col_idx in range(1, len(ext_headers) + 1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = 20

    # ── Sheet 3: Validation Errors ────────────────────────────────────────────
    ws3 = wb.create_sheet("Validation Errors")
    err_headers = ["File Name", "Field", "Message", "Severity"]
    for col_idx, h in enumerate(err_headers, start=1):
        cell = ws3.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    err_row = 2
    for result in results:
        for err in result.validation_errors:
            sev_fill = warning_fill if err.severity == "warning" else error_fill
            vals = [result.filename, err.field, err.message, err.severity]
            for col_idx, val in enumerate(vals, start=1):
                cell = ws3.cell(row=err_row, column=col_idx, value=val)
                cell.fill = sev_fill
                cell.border = thin_border
            err_row += 1

    # Save to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
